import pathlib
import os
import pyodbc
import time
import openpyxl
import pandas as pd
from io import BytesIO
import streamlit as st
from datetime import datetime
from  openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles.colors import Color
from db import connect_to_db

@st.cache_data
def fetch_age_sum(dealer_id):
    return pd.read_sql_query(
        """
        SELECT distinct
        Brand,Dealer,Location,SUM([NM 6Mths]) AS [NM 6Mths],SUM([NM 12 Mths]) AS [NM 12 Mths],SUM([NM 24 Mths]) AS [NM 24 Mths],
        SUM([NM 6Mths])+SUM([NM 12 Mths])+SUM([NM 24 Mths]) As [Total Non moving]
        FROM (SELECT distinct Brand,Dealer,Dealerid,Location,Ageing_Category,SUM([Stock Value(as per NDP Price)] )AS [Stock Value(as per NDP Price)],
            MAX(stockdate) AS stockdate,
            CASE WHEN Ageing_Category = 'NM 6 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 6Mths],
            CASE WHEN Ageing_Category = 'NM 12 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 12 Mths],
            CASE WHEN Ageing_Category = 'NM 24 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 24 Mths]
        FROM automation..AgeingAnalysisData_Automailer
        where Max_qty='0.00' AND Toc_qty='0.00'
        AND CAST(Addeddate AS date) = CAST(GETDATE() AS date)
        GROUP BY Brand,Dealer,Location,Dealerid,Ageing_Category) AS subquery
        where DEALERID=?
        GROUP BY Brand, Dealer,Location
        ORDER BY Brand,Dealer,SUM([NM 6Mths])+SUM([NM 12 Mths])+SUM([NM 24 Mths]) DESC;
        """,
        conn,
        params=(int(dealer_id),),
    )

conn = connect_to_db()
cursor = conn.cursor()

loc_df = pd.read_sql_query("select distinct a.Brand,a.Dealer,A.Brandid,A.Dealerid," \
"b.Dealer_Status as Category from locationinfo a " \
"INNER JOIN Dealer_Status_for_Ageing_Automation B on a.Dealerid = B.Dealerid ",conn)

st.set_page_config(page_title="Ageing Analysis App", layout="wide")

st.title("Ageing Analysis App")


# --- Sidebar Filters ---
st.sidebar.title("⚙ Settings")
st.sidebar.markdown("#### Select Category")

brand = st.sidebar.selectbox("Select Brand", loc_df["Brand"].unique())
category = st.sidebar.radio("Choose:", ["All", "Gainer", "Ogys"], horizontal=True)
#brandid = loc_df[loc_df["Brand"] == brand]["Brandid"].unique()
selected_brand_id = int(loc_df.loc[loc_df["Brand"] == brand, "Brandid"].iloc[0])

# ---- Persistent state (AFTER brand & category exist) ----
if "selected_dealer_ids" not in st.session_state:
    st.session_state.selected_dealer_ids = []

if "brand_snapshot" not in st.session_state:
    st.session_state.brand_snapshot = None
if "category_snapshot" not in st.session_state:
    st.session_state.category_snapshot = None

# Reset cached selections if brand/category changes
if st.session_state.brand_snapshot != brand or st.session_state.category_snapshot != category:
    st.session_state.selected_dealer_ids = []
    st.session_state.brand_snapshot = brand
    st.session_state.category_snapshot = category



# Filter dealer list based on selected category
#filtered_df = loc_df[loc_df["Brandid"].isin(brandid)]
filtered_df = loc_df[loc_df["Brandid"] == selected_brand_id]

if category == "Gainer":
    filtered_df = filtered_df[filtered_df["Category"] == "Gainer"]
elif category == "Ogys":
    filtered_df = filtered_df[filtered_df["Category"] == "Ogys"]

dealer_options = filtered_df["Dealer"].unique()
Dealer = st.sidebar.multiselect("Select Dealer(s)", options=dealer_options)
Dealer_id = loc_df[loc_df["Dealer"].isin(Dealer)]["Dealerid"].unique()

tab1, tab2, tab3, tab4= st.tabs(["Data Availability", "Dealer Status", "Summary", "Mail send"])

with tab1:
    selected_dealer_ids = []
    if Dealer_id is not None and len(Dealer_id) > 0:
        st.subheader("Dealer Data Status Check")
        if st.button("Check Status"):
            st.session_state.selected_dealer_ids = list(map(int, Dealer_id))
            #selected_dealer_ids.append(dealer_id)
            for dealer_id in  st.session_state.selected_dealer_ids:
                selected_dealer_ids.append(dealer_id)
                #dealer_id = int(dealer_id[0])
                st.write(f"### Dealer ID: {dealer_id}")
                brandid = loc_df.loc[loc_df["Dealerid"] == dealer_id, "Brandid"].values[0]
                dealer_name = loc_df.loc[loc_df["Dealerid"] == dealer_id, "Dealer"].values[0]
                brand_name = loc_df.loc[loc_df["Dealerid"] == dealer_id, "Brand"].values[0]
                brandid = str(brandid)


                st.write(f"### Dealer ID: {dealer_name} (Brand ID: {brand_name})")
                sale_pur_Df = pd.read_sql_query(f"""declare @cur_month as datetime,
                @Start int,
                @end int,
                @cs nvarchar(max),
                @p nvarchar(max) ,
                @ws nvarchar(max),
                @CS_comb as nvarchar(max),
                @ws_Comb as nvarchar(max) ,
                @p_comb as nvarchar(max),
                @dealerid as nvarchar(54)


                select @dealerid = bigid from Dealer_Master   where bigid= '{dealer_id}'

                set @cur_month = getdate()
                set @CS_comb =''
                set @ws_Comb =''
                set @p_comb =''


                set @Start=1
                set @end=24

                while @Start<=@end
                begin

                set @cs =concat(FORMAT(dateadd(month,-@start,@cur_month),'MMM_yy'),'_cs')
                set @ws =concat(FORMAT(dateadd(month,-@start,@cur_month),'MMM_yy'),'_ws')
                set @p = concat(FORMAT(dateadd(month,-@start,@cur_month),'MMM_yy'),'_P')

                set @CS_comb =@CS_comb+@cs+','
                set @ws_Comb = @ws_Comb+@ws+','
                set @p_comb = @p_comb+@p+','

                set @Start =@Start+1
                end

                set @CS_comb = left(@CS_comb,len(@cs_comb)-1)
                set @ws_Comb = LEFT(@ws_comb,len(@ws_comb)-1)
                set @p_comb  = left(@p_comb,len(@p_comb)-1)



                declare @sql nvarchar(max)

                Set @sql ='
                WITH DATA_CHECK1 AS(

                select brandid,dealerid,locationid,columnj,sum(cast(value as decimal)) as value,
                SUBSTRING(columnj,CHARINDEX(''_'',columnj,CHARINDEX(''_'',columnj)+1)+1,LEN(columnj)) AS Data_type,
                case when (case when right(columnj ,1)=''p''  then  sum(cast(value as decimal)) else 0 end)>0 then 0
                else 1 end	 Pur_status,
                case when (case when right(columnj ,3)=''_cs''  then  sum(cast(value as decimal)) else 0 end)>0 then 0
                else 1 end	 Cs_status,
                case when (case when right(columnj ,3)=''_ws''  then  sum(cast(value as decimal)) else 0 end)>0 then 1
                else 1 end	 Ws_status

                from(
                select brandid,dealerid,locationid,
                '+@CS_comb+' ,'+@ws_Comb+','+@p_comb+'
                from Dealer_Sale_Upload_Old_TD001_'+@dealerid+'
                ) as tbl

                unpivot(
                value for columnj  in ( '+@CS_comb+' ,'+@ws_Comb+','+@p_comb+')
                ) vn

                group by  brandid,dealerid,locationid,columnj	)

                ---------------- FINAL OUTPUT FOR DATA CHECK-----------------

                SELECT b.Brand,B.Dealer,b.Location,ConsigneeType,
                a.BRANDID,a.Dealerid,a.Locationid,Data_type,columnj as Month_of_data,value,
                --sum(value) as value,
                case when value > 0.0 then ''OK'' else ''Not ok'' end Data_check
                fROM  DATA_CHECK1 a
                inner join locationinfo b on a.locationid = b.locationid
                group by a.BRANDID,a.Dealerid,a.Locationid ,Data_type,columnj,value,b.Brand,B.Dealer,b.Location,ConsigneeType
                '
                exec sp_executesql @sql, N'@dealerid nvarchar(54)', @dealerid = {dealer_id}""", conn)
                cp_df = pd.read_sql_query(f"exec Cp_ageing_datacheck '{dealer_id}'",conn)
                #cursor.execute("exec indal_AgeingAnalysis_Automailer ?,?", (str(brandid), str(dealer_id)))
                #conn.commit()

                data = pd.concat([sale_pur_Df, cp_df], ignore_index=True)


                data['Data_check'] = data['Data_check'].str.lower()
                data['Data_type'] = data['Data_type'].str.lower()  

                filtered = data[(data['Data_check'] == 'ok')]
                filtered['Data_type'] = filtered['Data_type'].str.lower()
                
                df_c = filtered.pivot_table(index=['Brand', 'Dealer', 'Location','ConsigneeType'],columns='Data_type',values='Data_check',aggfunc='count').reindex(columns=['p', 'cs', 'ws', 'cp'], fill_value=0)
                df_c.reset_index(inplace =True)
                #st.data_editor(df_c)
                df_c['Pur_status']=df_c['p'].apply(lambda x:'Ok' if x==24 else 'Not ok')
                df_c['Cs_status']=df_c['cs'].apply(lambda x:'Ok' if x==24 else 'Not ok')
                df_c['Ws_status']=df_c['ws'].apply(lambda x:'Ok' if x==24 else 'Not ok')
                df_c['Cp_status']=df_c['cp'].apply(lambda x:'Ok' if x==24 else 'Not ok')

                #if (df_c[['Ws_status', 'Cs_status', 'Pur_status','Cp_status']] == 'Not ok').any(axis=None):
                status_cols = ['Ws_status', 'Cs_status', 'Pur_status', 'Cp_status']

                if (df_c[status_cols] == 'Not ok').any(axis=None):
                    dealer_name = loc_df.loc[loc_df["Dealerid"] == dealer_id, "Dealer"].values[0]
                    
                    # Choose emoji based on status
                    final_status_emoji = "⚠️"  # default warning
                    if (df_c[status_cols] == 'Not ok').any(axis=None):
                        final_status_emoji = "⚠️"  # warning
                    else:
                        final_status_emoji = "✅"  # all good

                dealer_name = loc_df[loc_df["Dealerid"] == dealer_id]["Dealer"].values[0]
                with st.expander(f"📁  {dealer_name} (ID: {dealer_id})",):
                    st.subheader("⚠️ Final Summary Status")
                    st.dataframe(df_c, use_container_width=True)
                    st.subheader("🔍 Raw Data Details")
                    st.dataframe(data, use_container_width=True)
                    col1, col2 = st.columns([1, 2])
                
                    # Download buttons
                    with col1:
                        st.download_button(
                            key=f"download_summary_{dealer_id}",
                            label="⬇️ Download Summary CSV",
                            data=df_c.to_csv(index=False),
                            file_name=f"{dealer_name}_summary.csv",
                            mime="text/csv"
                        )
                    with col2:
                        st.download_button(
                            key=f"download_raw_{dealer_id}",
                            label="⬇️ Download Raw Data CSV",
                            data=data.to_csv(index=False),
                        file_name=f"{dealer_name}_raw.csv",
                        mime="text/csv"
                    )

    with tab2:
        allowed = loc_df[loc_df["Dealerid"].isin(st.session_state.selected_dealer_ids)]["Dealer"].unique()
        selected_deale = st.multiselect("Select Dealer(s)", options=allowed, key="tab3_select_dealers")

        selected_ids_df = (
            loc_df[(loc_df["Dealer"].isin(selected_deale)) & (loc_df["Brand"] == brand)][["Brandid", "Dealerid"]]
            .dropna()
            .astype(int)
        )

        if st.button("Generate NonMoving Report") and not selected_ids_df.empty:
            for dealer_id, brand_id in zip(selected_ids_df["Dealerid"], selected_ids_df["Brandid"]):
                exec_sql = """
                SET NOCOUNT ON;
                EXEC dbo.indal_AgeingAnalysis_Automailer @BrandID = ?, @DealerID = ?;
                """
                params = (int(brand_id), int(dealer_id))
                st.write("Executing query...", params)

                with conn.cursor() as cur:
                    cur.execute(exec_sql, params)
                    conn.commit()

                df = pd.read_sql_query(
                    """
                    SELECT *
                    FROM automation..AgeingAnalysisData_Automailer
                    WHERE CAST(Addeddate AS date) = CAST(GETDATE() AS date)
                    AND DealerID = ?
                    AND BrandID = ?
                    """,
                    conn,
                    params=(int(dealer_id), int(brand_id)),
                )
                #st.dataframe(df.head(), use_container_width=True)
                if len(df)>0:
                    st.success(f"Data for Dealer ID {dealer_id} and Brand ID {brand_id} fetched successfully.")    

        else:
            st.info("Pick at least one dealer from the list you checked in Data Availability.")

    with tab3:
        st.subheader("Summary of Data Availability")
        #st.multiselect("Select Dealer(s)", options=loc_df["Dealer"].unique(), key="summary_select_dealers")
        allowed = loc_df[loc_df["Dealerid"].isin(st.session_state.selected_dealer_ids)]["Dealer"].unique()
        selected_deale = st.multiselect("Select Dealer(s)", options=allowed, key="tab3_select_dealer_s")

        selected_ids_df = (
            loc_df[(loc_df["Dealer"].isin(selected_deale)) & (loc_df["Brand"] == brand)][["Brandid", "Dealerid"]]
            .dropna()
            .astype(int))
        
        if st.button("Fetch Summary") and not selected_ids_df.empty:

            for dealer_id, brand_id in zip(selected_ids_df["Dealerid"], selected_ids_df["Brandid"]):

                Age_sum = fetch_age_sum(dealer_id)
                st.session_state[f"age_sum_{dealer_id}"] = Age_sum
                st.session_state[f"csv_{dealer_id}"] = Age_sum.to_csv(index=False)
                st.session_state[f"expander_{dealer_id}"] = True
                with st.expander(label=f"📁 Summary of Ageing Data for  {dealer_id}", expanded=st.session_state.get(f"expander_{dealer_id}", False)):
                    if len(Age_sum) > 0:
                        try:
                            st.download_button(
                                key=f"download_summary_{dealer_id}",
                                label=f"Download summary for {dealer_id}",
                                data=st.session_state[f"csv_{dealer_id}"],
                                file_name=f"Ageing summary for {dealer_id}.csv",
                                mime="text/csv"
                            )
                            st.header("Summary of Ageing Data for Dealer ID {}".format(dealer_id))
                            st.dataframe(Age_sum, use_container_width=True)
                        except:
                            st.warning(f"Something wrong with {dealer_id}")





        else:
            st.info("No dealers selected. Please check the Data Availability tab first.")

    with tab4:
        
        st.markdown("<h3 style='color: Green;'>Send Mail</h3>", unsafe_allow_html=True)
        if st.button("Click Here to send Mail",key="Mail send") and not selected_ids_df.empty:
            Mail_df =pd.read_sql_query('''select *from EmailAggregation_view_Ageing_Automailer ''',conn)
            Mail_df.rename(columns={'ToEmail':'To','CcEmail':'CC'},inplace=True)
            merge_df = selected_ids_df.merge(Mail_df, left_on='Dealerid', right_on='Dealerid', how='inner')
  

            for dealer_id, brand_id, to_email, cc_email, dealer_name in zip(merge_df["Dealerid"], merge_df["Brandid"], merge_df["To"], merge_df["CC"],merge_df["dealer"]):  
               
                sug_dis = pd.read_sql_query('select *from automation..uad_Suggesting_Discount_Master_Gainer where brandid=?',conn,params=int(brand_id),)   
                sql = f"""
                        SELECT A.*, C.OrderPartNumber
                        FROM automation..AgeingAnalysisData_Automailer AS a
                        LEFT JOIN Part_Master c
                            ON a.Brandid = c.brandid
                            AND REPLACE(A.PartNumber, '''', '') = REPLACE(c.partnumber, '''', '')
                        WHERE a.[Stock Qty] > 0 AND A.Brandid = {brand_id}
                        AND A.Dealerid = {dealer_id}
                        AND CAST(A.Addeddate AS date) = CAST(GETDATE() AS date)
                    """

                file_for = pd.read_sql_query(sql, conn)

                file_for['Purchase6M']=file_for['Purchase6M'].astype(float)+file_for['CODPurchase6M'].astype(float)
                file_for['Purchase12M']=file_for['Purchase12M'].astype(float)+file_for['CODPurchase12M'].astype(float)
                file_for['Purchase24M']=file_for['CODPurchase24M'].astype(float)+file_for['Purchase24M'].astype(float)    


                file_for['Sugg. DiscPercentage'] = file_for.apply(
                    lambda row: sug_dis['NM_6M_Disc'].values[0] if row['Ageing_Category'] == 'NM 6 Mths' and row['Brandid'] == sug_dis['Brandid'].values[0]
                    else sug_dis['NM_12M_Disc'].values[0] if row['Ageing_Category'] == 'NM 12 Mths' and row['Brandid'] == sug_dis['Brandid'].values[0]
                    else sug_dis['NM_24M_Disc'].values[0] if row['Ageing_Category'] == 'NM 24 Mths' and row['Brandid'] == sug_dis['Brandid'].values[0]
                    else None,
                    axis=1
                )


                to_excel = file_for[(file_for['Ageing_Category'].notnull())&(file_for['Max_qty'].astype(float)<1.0)&(file_for['Toc_qty'].astype(float)<1.0)][[
                        'Brand', 'Dealer', 'Location','PartNumber', 'LatestPartNumber','OrderPartNumber','Description',
                        'MRP', 'NDP', 'Stock Qty', 'Stock Value(as per NDP Price)', 'StockDate',
                        'Sale6M', 'Purchase6M', 'Sale12M', 'Purchase12M', 'Sale24M', 'Purchase24M',
                        'Ageing_Category','Sugg. DiscPercentage']]

                    # Ageing Summary for sent
                Age_sum = pd.read_sql_query(
                        """SELECT distinct
                        Brand,Dealer,Location,SUM([NM 6Mths]) AS [NM 6Mths],SUM([NM 12 Mths]) AS [NM 12 Mths],SUM([NM 24 Mths]) AS [NM 24 Mths],
                        SUM([NM 6Mths])+SUM([NM 12 Mths])+SUM([NM 24 Mths]) As [Total Non moving]
                        FROM (SELECT distinct Brand,Dealer,Dealerid,Location,Ageing_Category,SUM([Stock Value(as per NDP Price)] )AS [Stock Value(as per NDP Price)],
                            MAX(stockdate) AS stockdate,
                            CASE WHEN Ageing_Category = 'NM 6 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 6Mths],
                            CASE WHEN Ageing_Category = 'NM 12 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 12 Mths],
                            CASE WHEN Ageing_Category = 'NM 24 Mths' THEN SUM([Stock Value(as per NDP Price)]) ELSE 0 END AS [NM 24 Mths]
                            FROM automation..AgeingAnalysisData_Automailer
                        where Max_qty='0.00' AND Toc_qty='0.00'	AND CAST(Addeddate AS date) = CAST(GETDATE() AS date)
                        GROUP BY Brand,Dealer,Location,Dealerid,Ageing_Category) AS subquery
                        where DEALERID=?
                    GROUP BY Brand, Dealer,Location
                    having (SUM([NM 6Mths])+SUM([NM 12 Mths])+SUM([NM 24 Mths]))>0
                    ORDER BY Brand,Dealer,SUM([NM 6Mths])+SUM([NM 12 Mths])+SUM([NM 24 Mths]) DESC;
                        """,
                        conn,
                        params=(int(dealer_id),))
                 #MAil sending
                
                Document_folder = str(pathlib.Path.home() / "Documents")
            
                #filename
                output_excel =Document_folder+"\ Suggested_Non_Moving_Report_for_" +dealer_name +" "+datetime.strftime(datetime.now(),'%Y-%b')+ ".xlsx"
                if len(to_excel) > 0:
                    t = to_excel.iloc[0,0]
                    to_excel.to_excel(output_excel,index=False)
                    #excel formating
                    file_path = output_excel
                    wb = load_workbook(file_path)
                    ws = wb.active
                    ws.alignment = Alignment(horizontal='center', vertical='center')
                    for cell in ws[1]:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        cell.fill = cell.fill.copy(fgColor='a7a758')
                        cell.fill = PatternFill('solid', start_color="38ffe9")
                    max_row = ws.max_row
                    max_col = ws.max_column
                    for row in range(1, max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.border = Border(left=Side(border_style=BORDER_THIN, color='000000'),
                                                right=Side(border_style=BORDER_THIN, color='000000'),
                                                top=Side(border_style=BORDER_THIN, color='000000'),
                                                bottom=Side(border_style=BORDER_THIN, color='000000'))
                    wb.save(file_path)
                    # Add your email sending logic here
                    import smtplib as s
                    from email.mime.text import MIMEText
                    from email.mime.multipart import MIMEMultipart
                    from email.mime.base import MIMEBase
                    from email import encoders
                    import time

                    # SMTP server setup
                    ob = s.SMTP('smtp.gmail.com', 587)
                    ob.ehlo()
                    ob.starttls()
                    print("Server connected")
                    ob.login('gainer.alerts@sparecare.in','lbcjiqdlbhtnfouq')

                    # Email details
                    
                    # subject = 'Suggested Non Moving :- '+dealer_name
                    # body = f'''
                    # <!-- start mail body -->
                    # '''

                    # # ------- build HTML from DataFrames -------
                    # def fmt_money(x):
                    #     try:
                    #         return "{:,.0f}".format(float(x))
                    #     except Exception:
                    #         return "0"

                    # # Summary table rows from Age_sum (one row per Location + grand total)
                    # td_l = "border:1px solid #d1d5db;"
                    # td_r = "border:1px solid #d1d5db; text-align:right;"

                    # summary_rows = ""
                    # if not Age_sum.empty:
                    #     for _, r in Age_sum.iterrows():
                    #         summary_rows += f"""
                    #         <tr>
                    #         <td style="{td_l}">{r['Location']}</td>
                    #         <td style="{td_r}">{fmt_money(r['NM 6Mths'])}</td>
                    #         <td style="{td_r}">{fmt_money(r['NM 12 Mths'])}</td>
                    #         <td style="{td_r}">{fmt_money(r['NM 24 Mths'])}</td>
                    #         <td style="{td_r}">{fmt_money(r['Total Non moving'])}</td>
                    #         </tr>
                    #         """

                    #     gt6  = Age_sum['NM 6Mths'].sum()
                    #     gt12 = Age_sum['NM 12 Mths'].sum()
                    #     gt24 = Age_sum['NM 24 Mths'].sum()
                    #     gtot = Age_sum['Total Non moving'].sum()
                        
                    #     summary_rows += f"""
                    #     <tr>
                    #     <td style="{td_l};font-weight:bold;">Grand Total</td>
                    #     <td style="{td_r};font-weight:bold;">{fmt_money(gt6)}</td>
                    #     <td style="{td_r};font-weight:bold;">{fmt_money(gt12)}</td>
                    #     <td style="{td_r};font-weight:bold;">{fmt_money(gt24)}</td>
                    #     <td style="{td_r};font-weight:bold;">{fmt_money(gtot)}</td>
                    #     </tr>
                    #     """
                    # else:
                    #     gt6 = gt12 = gt24 = gtot = 0

                    # # Suggested discount table from sug_dis
                    # disc6  = sug_dis.iloc[0]['NM_6M_Disc']  if 'NM_6M_Disc'  in sug_dis.columns and not sug_dis.empty else 0
                    # disc12 = sug_dis.iloc[0]['NM_12M_Disc'] if 'NM_12M_Disc' in sug_dis.columns and not sug_dis.empty else 0
                    # disc24 = sug_dis.iloc[0]['NM_24M_Disc'] if 'NM_24M_Disc' in sug_dis.columns and not sug_dis.empty else 0

                    # disc_rows = f"""
                    #     <tr><td style="{td_l}">Non Moving – 6 Months</td><td style="border:1px solid #d1d5db; text-align:center;">{disc6}%</td></tr>
                    #     <tr><td style="{td_l}">Non Moving – 12 Months</td><td style="border:1px solid #d1d5db; text-align:center;">{disc12}%</td></tr>
                    #     <tr><td style="{td_l}">Non Moving – 24 Months</td><td style="border:1px solid #d1d5db; text-align:center;">{disc24}%</td></tr>
                    # """
                    # # ------------------------------------------
                    # #subject = f"Suggested Non Moving Report for {dealer_name} (Rs. {fmt_money(gtot)/10000:.2f} Lacs)"
                    # subject = f"Suggested Non Moving Report for {dealer_name} (Rs. {gtot/100000:.2f} Lacs)"
    
                    # #subject = 'Suggested Non Moving :- ' + dealer_name
                    # body = f"""
                    # <!-- start mail body -->
                    # <div style="font-family: Arial, Helvetica, sans-serif; font-size:14px; color:#1f2937; line-height:1.5;">
                    # <p style="margin:0 0 8px 0;">Dear Sir,</p>
                    # <p style="margin:0 0 16px 0;">Greetings !!</p>

                    # <p style="margin:0 0 10px 0;">
                    #     Kindly find attached the list of <b>Non Moving Spare Parts</b> suggested for listing on
                    #     <b>Sparecare Gainer Portal</b> for liquidation.
                    # </p>

                    # <!-- summary table (from Age_sum) -->
                    # <table role="presentation" cellspacing="0" cellpadding="6" border="0"
                    #         style="border-collapse:collapse; width:100%; max-width:720px; margin:8px 0; border:1px solid #d1d5db;">
                    #     <thead>
                    #     <tr>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:left;">Location</th>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:center;">Non Moving<br>06 Mths</th>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:center;">Non Moving<br>12 Mths</th>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:center;">Non Moving<br>24 Mths</th>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:center;">Grand Total</th>
                    #     </tr>
                    #     </thead>
                    #     <tbody>
                    #     {summary_rows}
                    #     </tbody>
                    # </table>

                    # <p style="margin:6px 0; font-size:12px; color:#374151;">
                    #     # NM = Non Moving<br>
                    #     ## Total Slow / Non Moving Parts Value in Rs.
                    #     <span style="background:#fff59d; padding:1px 4px;">{fmt_money(gtot)}</span> /-
                    # </p>

                    # <!-- criteria -->
                    # <p style="margin:14px 0 6px 0; font-weight:bold;">Criteria for Non Moving Parts Selection</p>
                    # <ol style="margin:0 0 12px 18px; padding:0;">
                    #     <li style="margin:4px 0;"><b>Non Moving 06 months</b> : Parts with “0” Sales &amp; Purchase in last 06 mths.</li>
                    #     <li style="margin:4px 0;"><b>Non Moving 12 months</b> : Parts with “0” Sales &amp; Purchase in last 12 mths.</li>
                    #     <li style="margin:4px 0;"><b>Non Moving 24 months</b> : Parts with “0” Sales &amp; Purchase in last 24 mths.</li>
                    # </ol>

                    # <!-- Suggested Discount (from sug_dis) -->
                    # <p style="margin:10px 0 6px 0; font-weight:bold;">Suggested Discount</p>
                    # <table role="presentation" cellspacing="0" cellpadding="6" border="0"
                    #         style="border-collapse:collapse; width:100%; max-width:520px; border:1px solid #d1d5db; margin:6px 0 12px 0;">
                    #     <thead>
                    #     <tr>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:left;">Parts Ageing</th>
                    #         <th style="background:#e5e7eb; border:1px solid #d1d5db; text-align:center;">Discount % on<br>Retail/Taxable value</th>
                    #     </tr>
                    #     </thead>
                    #     <tbody>
                    #     {disc_rows}
                    #     </tbody>
                    # </table>

                    # <p style="margin:6px 0;">Minimum Discount suggested to Dealer. A Dealer may give higher discount to increase chances of Liquidation</p>
                    # <p style="margin:0 0 10px 0; color:#b91c1c; font-weight:bold;">NOTE : Seller cost of carrying the dead stock is higher than proposed discount</p>

                    # <p style="margin:10px 0;">It is requested to kindly check the details and in case of any feedback, please revert.</p>
                    
                    # <p style="margin:0 0 10px 0;">In case of no reply received in next 4 day, Suggested Non Moving Parts list will be updated in GAINER portal for Liquidation.</p>
                    
                    # <p style="font-family:Calibri, Arial, Helvetica, sans-serif;">Thanks &amp; Regards<br>Team Gainer</p>
                    # </div>
                    # <!-- end mail body -->
                    # """

                    # ------- build HTML from DataFrames -------
                    def fmt_money(x):
                        try:
                            return "{:,.0f}".format(float(x))
                        except Exception:
                            return "0"

                    # --- mail table styles (match screenshot) ---
                    HDR_BG = "#cfeaf6"               # light blue header
                    BORDER = "1px solid #000000"     # black grid
                    PAD = "6px 8px"

                    th_l = f"border:{BORDER}; background:{HDR_BG}; text-align:left;  padding:{PAD};"
                    th_c = f"border:{BORDER}; background:{HDR_BG}; text-align:center; padding:{PAD};"

                    td_l = f"border:{BORDER}; padding:{PAD};"
                    td_r = f"border:{BORDER}; text-align:right; padding:{PAD};"

                    # Robust numeric coercion (columns as in your SQL)
                    for colname in ["NM 6 Mths", "NM 12 Mths", "NM 24 Mths", "Total Non moving"]:
                        if colname in Age_sum.columns:
                            Age_sum[colname] = pd.to_numeric(Age_sum[colname], errors="coerce").fillna(0)

                    # ---- build summary rows from Age_sum ----
                    summary_rows = ""
                    if not Age_sum.empty:
                        for _, r in Age_sum.iterrows():
                            summary_rows += f"""
                            <tr>
                            <td style="{td_l}">{r['Location']}</td>
                            <td style="{td_r}">{fmt_money(r['NM 6Mths'])}</td>
                            <td style="{td_r}">{fmt_money(r['NM 12 Mths'])}</td>
                            <td style="{td_r}">{fmt_money(r['NM 24 Mths'])}</td>
                            <td style="{td_r}">{fmt_money(r['Total Non moving'])}</td>
                            </tr>
                            """

                        gt6  = Age_sum["NM 6Mths"].sum()
                        gt12 = Age_sum["NM 12 Mths"].sum()
                        gt24 = Age_sum["NM 24 Mths"].sum()
                        gtot = Age_sum["Total Non moving"].sum()

                        # highlighted last row (Grand Total)
                        summary_rows += f"""
                        <tr style="background:#e6f7ff;">
                        <td style="{td_l}font-weight:bold;">Grand Total</td>
                        <td style="{td_r}font-weight:bold;">{fmt_money(gt6)}</td>
                        <td style="{td_r}font-weight:bold;">{fmt_money(gt12)}</td>
                        <td style="{td_r}font-weight:bold;">{fmt_money(gt24)}</td>
                        <td style="{td_r}font-weight:bold;">{fmt_money(gtot)}</td>
                        </tr>
                        """
                    else:
                        gt6 = gt12 = gt24 = gtot = 0

                    # ---- full summary table html ----
                    summary_table_html = f"""
                    <table role="presentation" cellspacing="0" cellpadding="0" border="0"
                        style="border-collapse:collapse; width:100%; max-width:720px;">
                    <thead>
                        <tr>
                        <th style="{th_l}">Location</th>
                        <th style="{th_c}">Non Moving<br>06 Mths</th>
                        <th style="{th_c}">Non Moving<br>12 Mths</th>
                        <th style="{th_c}">Non Moving<br>24 Mths</th>
                        <th style="{th_c}">Grand Total</th>
                        </tr>
                    </thead>
                    <tbody>
                        {summary_rows}
                    </tbody>
                    </table>
                    """

                    # ---- Suggested Discount table from sug_dis ----
                    disc6  = sug_dis.iloc[0]["NM_6M_Disc"]  if ("NM_6M_Disc"  in sug_dis.columns and not sug_dis.empty) else 0
                    disc12 = sug_dis.iloc[0]["NM_12M_Disc"] if ("NM_12M_Disc" in sug_dis.columns and not sug_dis.empty) else 0
                    disc24 = sug_dis.iloc[0]["NM_24M_Disc"] if ("NM_24M_Disc" in sug_dis.columns and not sug_dis.empty) else 0

                    disc_rows = f"""
                    <tr><td style="{td_l}">Non Moving – 6 Months</td><td style="border:{BORDER}; text-align:center; padding:{PAD};">{disc6}%</td></tr>
                    <tr><td style="{td_l}">Non Moving – 12 Months</td><td style="border:{BORDER}; text-align:center; padding:{PAD};">{disc12}%</td></tr>
                    <tr><td style="{td_l}">Non Moving – 24 Months</td><td style="border:{BORDER}; text-align:center; padding:{PAD};">{disc24}%</td></tr>
                    """

                    disc_table_html = f"""
                    <table role="presentation" cellspacing="0" cellpadding="0" border="0"
                        style="border-collapse:collapse; width:100%; max-width:520px; margin-top:6px;">
                    <thead>
                        <tr>
                        <th style="{th_l}">Parts Ageing</th>
                        <th style="{th_c}">Discount % on<br>Retail/Taxable value</th>
                        </tr>
                    </thead>
                    <tbody>
                        {disc_rows}
                    </tbody>
                    </table>
                    """

                    # ---- Subject + Body (insert generated tables) ----
                    subject = f"Suggested Non Moving Report for {dealer_name} (Rs. {gtot/100000:.2f} Lacs)"

                    body = f"""
                    <!-- start mail body -->
                    <div style="font-family: Arial, Helvetica, sans-serif; font-size:14px; color:#1f2937; line-height:1.5;">
                    <p style="margin:0 0 8px 0;">Dear Sir,</p>
                    <p style="margin:0 0 16px 0;">Greetings !!</p>

                    <p style="margin:0 0 10px 0;">
                        Kindly find attached the list of <b>Non Moving Spare Parts</b> suggested for listing on
                        <b>Sparecare Gainer Portal</b> for liquidation.
                    </p>

                    {summary_table_html}

                    <p style="margin:6px 0; font-size:12px; color:#374151;">
                        # NM = Non Moving<br>
                        ## Total Slow / Non Moving Parts Value in Rs.
                        <span style="background:#fff59d; padding:1px 4px;">{fmt_money(gtot)}</span> /-
                    </p>

                    <p style="margin:14px 0 6px 0; font-weight:bold;">Criteria for Non Moving Parts Selection</p>
                    <ol style="margin:0 0 12px 18px; padding:0;">
                        <li style="margin:4px 0;"><b>Non Moving 06 months</b> : Parts with “0” Sales &amp; Purchase in last 06 mths.</li>
                        <li style="margin:4px 0;"><b>Non Moving 12 months</b> : Parts with “0” Sales &amp; Purchase in last 12 mths.</li>
                        <li style="margin:4px 0;"><b>Non Moving 24 months</b> : Parts with “0” Sales &amp; Purchase in last 24 mths.</li>
                    </ol>

                    <p style="margin:10px 0 6px 0; font-weight:bold;">Suggested Discount</p>
                    {disc_table_html}

                    <p style="margin:6px 0;">Minimum Discount suggested to Dealer. A Dealer may give higher discount to increase chances of Liquidation</p>
                    <p style="margin:0 0 10px 0; color:#b91c1c; font-weight:bold;">NOTE : Seller cost of carrying the dead stock is higher than proposed discount</p>

                    <p style="margin:10px 0;">It is requested to kindly check the details and in case of any feedback, please revert.</p>
                    <p style="margin:10px 0;">In case of no reply received in next 4 day, Suggested Non Moving Parts list will be updated in GAINER portal for Liquidation.</P>
                    <p style="margin:15px 0;">Thanks &amp; Regards<br>Team Gainer</p>
                    
                    </div>
                    <!-- end mail body -->

                    """




                    #to_email=to
                    #cc_emails=cc
                    to_email ='idas98728@gmail.com'
                    cc_emails = 'scsit.db2@gmail.com'

                    #cc_emails = cc.replace(' ', '')
                    cc_email_list = cc_emails.split(',') if cc_emails else []
                    cc_email_list = [email for email in cc_email_list if email]  
                    
                    all_recipients = [to_email] + cc_email_list
                    
                    # Create a multipart message
                    msg = MIMEMultipart()
                    msg['From'] = 'gainer.alerts@sparecare.in'
                    msg['To'] = to_email
                    msg['Cc'] = ', '.join(cc_email_list)
                    msg['Subject'] = subject

                    msg.attach(MIMEText(body, 'html'))

                    # Open the attachment file and attach it to the email
                    filename = file_path
                    with open(filename, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f"attachment; filename= {filename.split('/')[-1]}")
                        msg.attach(part)

                    # Convert the message to a string and send it
                    text = msg.as_string()
                    ob.sendmail('gainer.alerts@sparecare.in', all_recipients, text)
                    time.sleep(5)
                    ob.quit()
                    print("Email sent successfully!")
                    st.success("Report sent successfully!" + dealer_name)
                   # st.success(f'Report sent successfully!: brand :{brd},Dealer :{dlrn},Location : {locn}')

                else:
                    st.warning('Report Not Genarate for :-'+dealer_name)
